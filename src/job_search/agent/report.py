from __future__ import annotations

from pathlib import Path
from typing import Any

from job_search.config import DIMENSIONS
from job_search.io_utils import read_jsonl

SHEET = "Matches"

# Wide enough to read at a glance; reasoning gets the remainder.
_WIDTHS: dict[str, int] = {
    "saved_at": 20,
    "score": 8,
    "reviewed": 10,
    "applied": 9,
    "title": 42,
    "company": 26,
    "location": 26,
    "job_url": 34,
    "overall_reasoning": 100,
}
_DIM_WIDTH = 14


def _row(rec: dict[str, Any], status: dict[str, Any]) -> dict[str, Any]:
    job = rec["job"]
    ev = rec["evaluation"]
    st = status.get(job["job_url"], {})
    scores = {d["name"]: d["score"] for d in ev["dimensions"]}

    row: dict[str, Any] = {
        "saved_at": rec["saved_at"],
        "score": ev["total"],
        "reviewed": bool(st.get("reviewed", False)),
        "applied": bool(st.get("applied", False)),
        "title": job["title"],
        "company": job["company"],
        "location": job["location"],
        "job_url": job["job_url"],
    }
    # The FitEvaluation validator guarantees all five dimensions in DIMENSIONS order,
    # so these lookups cannot miss.
    for dim in DIMENSIONS:
        row[dim.name] = scores[dim.name]
    row["overall_reasoning"] = ev["overall_reasoning"]
    return row


def build_xlsx(evaluations_path: Path, status: dict[str, Any], dest: Path) -> int:
    """Rebuild the Excel report in full from evaluations.jsonl + the reviewed/applied ticks.

    Regenerated from scratch every run rather than appended to, so it can never drift from
    the JSONL that is the source of truth, and a crashed run can't leave a half-written
    workbook behind. Newest first, best-scoring first within a day.

    Per-dimension *reasoning* deliberately stays in the JSONL — five paragraphs of prose
    per row make a spreadsheet unreadable. The scores come across as numbers you can sort
    and filter on.
    """
    import pandas as pd

    records = list(read_jsonl(evaluations_path))
    rows = [_row(rec, status) for rec in records]
    columns = [
        "saved_at", "score", "reviewed", "applied",
        "title", "company", "location", "job_url",
        *[d.name for d in DIMENSIONS],
        "overall_reasoning",
    ]
    df = pd.DataFrame(rows, columns=columns)
    if not df.empty:
        df = df.sort_values(["saved_at", "score"], ascending=[False, False])

    dest.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(dest, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=SHEET, index=False)
        _style(writer.sheets[SHEET], df, columns)
    return len(df)


def _style(sheet: Any, df: Any, columns: list[str]) -> None:
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    sheet.freeze_panes = "A2"
    for i, name in enumerate(columns, start=1):
        letter = get_column_letter(i)
        sheet.column_dimensions[letter].width = _WIDTHS.get(name, _DIM_WIDTH)
        sheet[f"{letter}1"].font = Font(bold=True)

    url_col = columns.index("job_url") + 1
    reasoning_col = columns.index("overall_reasoning") + 1
    for r in range(2, len(df) + 2):
        cell = sheet.cell(row=r, column=url_col)
        if cell.value:
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"
        sheet.cell(row=r, column=reasoning_col).alignment = Alignment(wrap_text=True, vertical="top")
